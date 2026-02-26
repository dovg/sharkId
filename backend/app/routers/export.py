"""
Excel export endpoints.

All endpoints require editor+ role and return .xlsx files as streaming responses.
"""
import io
from datetime import datetime, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.auth.dependencies import require_editor
from app.database import get_db
from app.models.dive_session import DiveSession
from app.models.location import Location
from app.models.observation import Observation
from app.models.photo import Photo, ProcessingStatus
from app.models.shark import Shark
from app.models.user import User
from app.utils.photo import photo_url

router = APIRouter(tags=["export"])

# ── helpers ───────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LINK_FONT   = Font(color="2D7DD2", underline="single")


def _make_wb(sheet_title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def _style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _hyperlink_cell(ws, row: int, col: int, url: str, label: str = "Open") -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = label
    cell.hyperlink = url
    cell.font = _LINK_FONT


def _fmt_dt(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _location_name(db: Session, location_id: Optional[UUID]) -> str:
    if location_id is None:
        return ""
    loc = db.get(Location, location_id)
    return f"{loc.spot_name}, {loc.country}" if loc else ""


# ── shark catalog ─────────────────────────────────────────────────────────────

@router.get("/sharks/export")
def export_sharks(
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Export the full shark catalog as Excel."""
    sharks = db.query(Shark).order_by(Shark.created_at).all()

    # Precompute first/last seen per shark from observations
    obs_all = db.query(Observation).filter(Observation.shark_id.isnot(None)).all()
    first_seen: dict[UUID, datetime] = {}
    last_seen:  dict[UUID, datetime] = {}
    for obs in obs_all:
        sid = obs.shark_id
        if obs.taken_at:
            if sid not in first_seen or obs.taken_at < first_seen[sid]:
                first_seen[sid] = obs.taken_at
            if sid not in last_seen or obs.taken_at > last_seen[sid]:
                last_seen[sid] = obs.taken_at

    obs_count: dict[UUID, int] = {}
    for obs in obs_all:
        obs_count[obs.shark_id] = obs_count.get(obs.shark_id, 0) + 1

    wb, ws = _make_wb("Shark Catalog")
    headers = ["Name", "Status", "First Seen", "Last Seen", "Observations", "Added", "Main Photo"]
    _style_header(ws, headers)

    for i, shark in enumerate(sharks, 2):
        main_url = None
        if shark.main_photo_id:
            mp = db.get(Photo, shark.main_photo_id)
            if mp:
                main_url = photo_url(mp)

        ws.cell(i, 1, shark.display_name)
        ws.cell(i, 2, shark.name_status.value)
        ws.cell(i, 3, _fmt_dt(first_seen.get(shark.id)))
        ws.cell(i, 4, _fmt_dt(last_seen.get(shark.id)))
        ws.cell(i, 5, obs_count.get(shark.id, 0))
        ws.cell(i, 6, _fmt_dt(shark.created_at))
        if main_url:
            _hyperlink_cell(ws, i, 7, main_url, "Photo")
        else:
            ws.cell(i, 7, "")

    _autofit(ws)
    return _xlsx_response(wb, "sharks.xlsx")


# ── shark detail ──────────────────────────────────────────────────────────────

@router.get("/sharks/{shark_id}/export")
def export_shark_detail(
    shark_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Export a single shark's observations with linked photos and GPS data."""
    shark = db.get(Shark, shark_id)
    if not shark:
        raise HTTPException(status_code=404, detail="Shark not found")

    observations = (
        db.query(Observation)
        .filter(Observation.shark_id == shark_id)
        .order_by(Observation.taken_at)
        .all()
    )

    wb, ws = _make_wb(shark.display_name[:31])  # sheet name max 31 chars
    headers = [
        "Date", "Location", "Session ID", "Comment", "Confirmed",
        "Photo", "Photo Taken At", "GPS Lat", "GPS Lon",
    ]
    _style_header(ws, headers)

    for i, obs in enumerate(observations, 2):
        loc_name = _location_name(db, obs.location_id)
        confirmed = "Yes" if obs.confirmed_at else "No"

        photo: Optional[Photo] = db.get(Photo, obs.photo_id) if obs.photo_id else None
        url = photo_url(photo) if photo else None
        taken_at = _fmt_dt(photo.taken_at if photo else None)
        gps_lat = photo.gps_lat if photo else None
        gps_lon = photo.gps_lon if photo else None

        ws.cell(i, 1, _fmt_dt(obs.taken_at))
        ws.cell(i, 2, loc_name)
        ws.cell(i, 3, str(obs.dive_session_id) if obs.dive_session_id else "")
        ws.cell(i, 4, obs.comment or "")
        ws.cell(i, 5, confirmed)
        if url:
            _hyperlink_cell(ws, i, 6, url, "Photo")
        else:
            ws.cell(i, 6, "")
        ws.cell(i, 7, taken_at)
        ws.cell(i, 8, gps_lat or "")
        ws.cell(i, 9, gps_lon or "")

    _autofit(ws)
    safe_name = "".join(c if c.isalnum() else "_" for c in shark.display_name)
    return _xlsx_response(wb, f"shark_{safe_name}.xlsx")


# ── sessions list ─────────────────────────────────────────────────────────────

@router.get("/dive-sessions/export")
def export_sessions(
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Export the full dive sessions list as Excel."""
    sessions = db.query(DiveSession).order_by(DiveSession.started_at.desc()).all()

    # Counts per session
    photos_q = (
        db.query(Photo.dive_session_id, Photo.id, Photo.processing_status, Photo.shark_id)
        .filter(Photo.dive_session_id.isnot(None))
        .all()
    )
    photo_count:  dict[UUID, int] = {}
    queue_count:  dict[UUID, int] = {}
    shark_ids_per_session: dict[UUID, set] = {}
    for p in photos_q:
        sid = p.dive_session_id
        photo_count[sid] = photo_count.get(sid, 0) + 1
        if p.processing_status == ProcessingStatus.ready_for_validation:
            queue_count[sid] = queue_count.get(sid, 0) + 1
        if p.shark_id:
            shark_ids_per_session.setdefault(sid, set()).add(p.shark_id)

    wb, ws = _make_wb("Dive Sessions")
    headers = ["Started", "Ended", "Location", "Comment", "Photos", "Queue", "Sharks"]
    _style_header(ws, headers)

    for i, session in enumerate(sessions, 2):
        loc_name = _location_name(db, session.location_id)
        sid = session.id
        ws.cell(i, 1, _fmt_dt(session.started_at))
        ws.cell(i, 2, _fmt_dt(session.ended_at))
        ws.cell(i, 3, loc_name)
        ws.cell(i, 4, session.comment or "")
        ws.cell(i, 5, photo_count.get(sid, 0))
        ws.cell(i, 6, queue_count.get(sid, 0))
        ws.cell(i, 7, len(shark_ids_per_session.get(sid, set())))

    _autofit(ws)
    return _xlsx_response(wb, "dive_sessions.xlsx")


# ── session detail ────────────────────────────────────────────────────────────

@router.get("/dive-sessions/{session_id}/export")
def export_session_detail(
    session_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Export all photos from a single dive session."""
    session = db.get(DiveSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Dive session not found")

    photos = (
        db.query(Photo)
        .filter(Photo.dive_session_id == session_id)
        .order_by(Photo.uploaded_at)
        .all()
    )

    # Observation lookup by photo_id
    obs_by_photo: dict[UUID, Observation] = {}
    for obs in db.query(Observation).filter(Observation.dive_session_id == session_id).all():
        if obs.photo_id:
            obs_by_photo[obs.photo_id] = obs

    wb, ws = _make_wb("Photos")
    headers = [
        "Photo", "Status", "Shark", "Taken At",
        "GPS Lat", "GPS Lon", "Orientation", "Observation",
    ]
    _style_header(ws, headers)

    for i, photo in enumerate(photos, 2):
        url = photo_url(photo)
        shark_name = ""
        if photo.shark_id:
            s = db.get(Shark, photo.shark_id)
            shark_name = s.display_name if s else ""

        obs = obs_by_photo.get(photo.id)
        obs_confirmed = ""
        if obs:
            obs_confirmed = "Confirmed" if obs.confirmed_at else "Draft"

        if url:
            _hyperlink_cell(ws, i, 1, url, "Photo")
        else:
            ws.cell(i, 1, "")
        ws.cell(i, 2, photo.processing_status.value)
        ws.cell(i, 3, shark_name)
        ws.cell(i, 4, _fmt_dt(photo.taken_at))
        ws.cell(i, 5, photo.gps_lat or "")
        ws.cell(i, 6, photo.gps_lon or "")
        ws.cell(i, 7, photo.orientation or "")
        ws.cell(i, 8, obs_confirmed)

    _autofit(ws)
    date_str = session.started_at.strftime("%Y-%m-%d")
    return _xlsx_response(wb, f"session_{date_str}.xlsx")
